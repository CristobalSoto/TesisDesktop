from openpyxl import load_workbook
from SQLServer import cnx
import gmpy2
import time
import warnings
import os
from sys import exit
warnings.filterwarnings("ignore")


# consulta general
def makequery(month, product, variety, quality, origin, unit, sectors):
    cueri = ("SELECT fecha_inicio, fecha_termino, producto, variedad, calidad, procedencia, tipo_producto, unidad, "
             "precio_promedio FROM [dbo].[ODEPA_Precios] "
             "WHERE (fecha_inicio LIKE '%" + month + "' OR fecha_termino LIKE '%" + month + "') "
             "AND producto in ('" + product + "') AND variedad " + variety + " AND calidad " + quality + " "
             "AND procedencia " + origin + " AND unidad ='" + unit + "' "
             "AND sector in(" + sectors + ") order by id")
    return cueri


# consulta para bebidas
def querybebidas(month, prod):
    cueri = ("SELECT fecha_inicio, fecha_termino, producto, variedad, calidad, procedencia, tipo_producto, unidad, "
             "precio_promedio FROM [dbo].[ODEPA_Precios] "
             "WHERE (fecha_inicio LIKE '%" + month + "' OR fecha_termino LIKE '%" + month + "') "
              "AND tipo_producto = '"+prod+"'")
    return cueri

time_start = time.time()
cursor = cnx.cursor()

# Obtiene directorio en el que se encuentra el programa y agrega el archivo al path
# file_name = "\\\seco\\scan$\\CPI MODEL\\00 - CL_CPI_Food_ODEPA.xlsm"
file_name = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))+ "\\00 - CL_CPI_Food_ODEPA.xlsm"
try:
    wb = load_workbook(file_name, keep_vba=True)
except FileNotFoundError:
    print("No se ha encontrado el archivo.")
    print("Recuerde que el nombre del archivo debe ser '01 - CL_CPI_Model.xlsm' y debe ubicarse en la carpeta Proyecto.")
    input("Presione ENTER para finalizar ")
    exit()

wsparams = wb['Parámetros CPI Food']
wsdicc = wb['Diccionario ODEPA - INE']
wsresumen = wb['ODEPA - CPI Food']

blank_spaces = 22
row_dicc = 3
row_precios = 3
col_ine = 8
col_odepa = 2
dicc_productos = {}
dicc_odepaine = {}
dicc_agrupaINE = {}
ine_list = []
lista_precios = []
sectores = "'"
lista_productos = []
count = 0
fila = 4
llave = ""
estrin = ""
prom_geo_hist = 1
n_ponderacion = 0
print("Obteniendo los datos desde el diccionario.....")
# Obtiene los sectores de las regiones que coinciden con las del INE
# Formatea el string para realizar la consulta en la base de datos
print("Obteniendo los sectores...")
while wsdicc.cell(row=fila, column=13).value is not None:
    if wsdicc.cell(row=fila, column=14).value is not None:
        sectores += wsdicc.cell(row=fila, column=13).value + "','"
    fila += 1

sectores = sectores[:-2]
# Obtiene los productos a buscar desde la hoja de diccionario
# Crea lista con productos de ine para ordenarlos al escribirlos en la hoja
print("Obteniendo los productos, variedades, procedencia y su unidad...")
while wsdicc.cell(row=row_dicc, column=2).value is not None:
    if wsdicc.cell(row=row_dicc, column=9).value is not None:
        tipo_producto = wsdicc.cell(row=row_dicc, column=col_odepa - 1).value
        producto = wsdicc.cell(row=row_dicc, column=col_odepa).value
        producto_ine = wsdicc.cell(row=row_dicc, column=col_ine).value
        variedad = wsdicc.cell(row=row_dicc, column=col_odepa + 1).value
        calidad = wsdicc.cell(row=row_dicc, column=col_odepa + 2).value
        # parsea el nulo
        if producto_ine == 'CARNE DE PAVO' and wsdicc.cell(row=row_dicc, column=col_odepa + 3).value is None:
            procedencia = ''
        else:
            procedencia = wsdicc.cell(row=row_dicc, column=col_odepa + 3).value
        unidad = wsdicc.cell(row=row_dicc, column=col_odepa + 4).value
        dicc_odepaine[producto_ine] = []
        if producto_ine not in ine_list:
            ine_list.append(producto_ine)
        row_precios += 1
        count += 1
        if producto not in lista_productos:
            lista_productos.append(producto)
        dicc_productos[producto_ine] = []

        dicc_productos[producto_ine] = producto, variedad, calidad, procedencia, unidad,
        try:
            if producto not in dicc_agrupaINE[producto_ine][0]:
                dicc_agrupaINE[producto_ine][0].append(producto)
            if variedad not in dicc_agrupaINE[producto_ine][1]:
                dicc_agrupaINE[producto_ine][1].append(variedad)
            if calidad not in dicc_agrupaINE[producto_ine][2]:
                dicc_agrupaINE[producto_ine][2].append(calidad)
            if procedencia not in dicc_agrupaINE[producto_ine][3]:
                dicc_agrupaINE[producto_ine][3].append(procedencia)
            if unidad not in dicc_agrupaINE[producto_ine][4]:
                dicc_agrupaINE[producto_ine][4].append(unidad)
        except KeyError:
            dicc_agrupaINE[producto_ine] = [[] for i in range(5)]
            if producto not in dicc_agrupaINE[producto_ine][0]:
                dicc_agrupaINE[producto_ine][0].append(producto)
            if variedad not in dicc_agrupaINE[producto_ine][1]:
                dicc_agrupaINE[producto_ine][1].append(variedad)
            if calidad not in dicc_agrupaINE[producto_ine][2]:
                dicc_agrupaINE[producto_ine][2].append(calidad)
            if procedencia not in dicc_agrupaINE[producto_ine][3]:
                dicc_agrupaINE[producto_ine][3].append(procedencia)
            if unidad not in dicc_agrupaINE[producto_ine][4]:
                dicc_agrupaINE[producto_ine][4].append(unidad)
    row_dicc += 1

col_busca = 9
sin_dato = False
# Busca la oolumna sin datos y registra la posicion de esta
print("Buscando la columna vacía....")
while True:
    for row in range(9, 43):
        if wsresumen.cell(row=row, column=col_busca).value is None:
            pass
        else:
            break
        if row == 42:
            sin_dato = True

    if sin_dato:
        break
    col_busca += 1

col_ine = 8
column = col_busca
print("Se actualizará desde la columna con la siguiente fecha:")
try:
    print("Fecha a actualizar: " + wsresumen.cell(row=8, column=column).value.strftime('%m/%Y'))
except AttributeError:
    print("Fecha a actualizar: " + wsresumen.cell(row=8, column=column - 1).value.strftime('%m/%Y'))
input("Presione ENTER para continuar")
inicio = False
print("Realizando las consultas correspondientes a la base de datos")
# Realiza la consulta en la BD y llena los campos de las columnas hasta que encuentre la columna final de la hoja
while True:
    for key in dicc_odepaine:
        dicc_odepaine[key] = []
    row = 9
    # si encuentra none en la columna de excel
    # se retrocede una para hacer la última sin importar si tiene datos
    try:
        mes = wsresumen.cell(row=8, column=column).value.strftime('%m/%Y')
    except AttributeError:
        column -= 1
        mes = wsresumen.cell(row=8, column=column).value.strftime('%m/%Y')

    for ine_product in ine_list:
        # genera los datos para realizar la consulta en la base de datos
        # transformando los datos a nulo si es necesario
        # o concatenandolos como tupla para la consulta "IN" 
        producto = "','".join(dicc_agrupaINE[ine_product][0])
        if dicc_agrupaINE[ine_product][1][0] is not None:
            estrin = ""
            for item in dicc_agrupaINE[ine_product][1]:
                if item == len(dicc_agrupaINE[ine_product][1]) - 1:
                    estrin += item.replace("'", "''")
                else:
                    estrin += item.replace("'", "''") + "','"
            variedad = "in ('" + estrin + "')"
        else:
            variedad = "is NULL"
        if dicc_agrupaINE[ine_product][2][0] is not None:
            calidad = "in ('" + "','".join(dicc_agrupaINE[ine_product][2]) + "')"
        else:
            calidad = "is NULL"
        if dicc_agrupaINE[ine_product][3][0] is not None:
            procedencia = "in ('" + "','".join(dicc_agrupaINE[ine_product][3]) + "')"
        else:
            procedencia = "is NULL"
        unidad = dicc_agrupaINE[ine_product][4][0]
        if producto == 'Bebidas':
            query = querybebidas(mes, producto)
        else:
            query = makequery(mes, producto, variedad, calidad, procedencia, unidad, sectores)
        cursor.execute(query)
        n_ponderacion = 0
        lista_precios = []
        # Consulta en la BD y definición de los dias del mes en base a la fecha de inicio y término
        prom_geo_hist = 1

        rowcount = 0
        for (fecha_inicio, fecha_termino, producto, variedad, calidad, procedencia, tipo_producto, unidad,
                precio_promedio) in cursor:
            produc_type = tipo_producto
            rowcount += 1
            fecha_ini_list = fecha_inicio.split('/')
            fecha_fin_list = fecha_termino.split('/')
            # realiza la determinación de la ponderacion por semana
            # compara los meses con esto realiza el cálculo correspondiente según el día para ponderar
            if fecha_ini_list[1] != fecha_fin_list[1]:
                if mes[:2] == fecha_ini_list[1]:
                    ponderacion = 5 - int(fecha_fin_list[0])
                    n_ponderacion += ponderacion
                    lista_precios.append(precio_promedio ** ponderacion)
                else:
                    ponderacion = int(fecha_fin_list[0])
                    n_ponderacion += ponderacion
                    lista_precios.append(precio_promedio ** ponderacion)
            else:
                ponderacion = 5
                lista_precios.append(precio_promedio**ponderacion)
                n_ponderacion += ponderacion

        print("Producto: " + ine_product, end="")
        for i in range(blank_spaces-len(ine_product)):
            print(" ",end="")
        for price in lista_precios:
            prom_geo_hist *= price
        try:
            prom_geo_hist = gmpy2.root(prom_geo_hist, n_ponderacion)
        except (ValueError, ZeroDivisionError):
            prom_geo_hist = None
        
        wsresumen.cell(row=row, column=column).value = float(prom_geo_hist) if prom_geo_hist is not None else prom_geo_hist
        try:
            print("Promedio geométrico: %d" % prom_geo_hist)
        except TypeError:
            print("Promedio geométrico: 0")
        row += 1
    column += 1
    if wsresumen.cell(row=8, column=column).value is None:
        break

print("Guardando los cambios en el documento excel.")
wb.save(file_name)
print("Cerrando la conexión a la base de datos.")
cursor.close()
cnx.close()

duracion = time.time() - time_start
print("Duración del programa: %s segundos" % duracion)
input("Ha finalizado de actualizar los productos, presione ENTER para finalizar el programa.")


