from openpyxl import load_workbook
from SQLServer import cnx
import os
import warnings
import sys


def makequery(table):
    cueri = ("SELECT DISTINCT region,sector from "+table+" WHERE tipo_producto !='Bebidas' order by region")
    return cueri
# diccionario de sectores en base de datos y en el excel
dicc_sectoresbd = {}
dicc_sectoreswb = {}
tabla = "[dbo].[ODEPA_Precios]"
warnings.filterwarnings("ignore")

print("Consultando a la base de datos.")
cursor = cnx.cursor()
query = makequery(tabla)
cursor.execute(query)
# guarda los sectores de la BD
for (region, sector) in cursor:
    try:
        dicc_sectoresbd[region].append(sector)
    except KeyError:
        dicc_sectoresbd[region] = []
        dicc_sectoresbd[region].append(sector)

print("Cargando la hoja Diccionario del documento excel.")
file_name = '\\\seco\\scan$\\CPI MODEL\\00 - CL_CPI_Food_ODEPA.xlsm'
try:
    wb = load_workbook(file_name, keep_vba=True)
except FileNotFoundError:
    print("No se ha encontrado el archivo.")
    print("Recuerde que el nombre del archivo debe ser '01 - CL_CPI_Model.xlsm' y debe ubicarse en la carpeta Proyecto.")
    input("Presione ENTER para finalizar ")
    sys.exit()
ws = wb['Diccionario ODEPA - INE']
row = 4
column = 12
# guarda los sectores del documento excel
while ws.cell(row=row, column=column+1).value is not None:
    if ws.cell(row=row, column=column).value is None:
        pass
    else:
        region = ws.cell(row=row, column=column).value
    sector = ws.cell(row=row, column=column+1).value
    try:
        dicc_sectoreswb[region].append(sector)
    except KeyError:
        dicc_sectoreswb[region] = []
        dicc_sectoreswb[region].append(sector)
    row += 1
# compara los sectores de ambos diccionarios
count = 0
for key in dicc_sectoresbd:
    for item in dicc_sectoresbd[key]:
        try:
            if item in dicc_sectoreswb[key]:
                pass
            else:
                print("Se ha encontrado un sector faltante")
                print("Region: "+key)
                print("Sector: "+item)
                input("Presione ENTER para agregar el sector al documento")
                ws.cell(row=row, column=column).value = key
                ws.cell(row=row, column=column+1).value = item
                row += 1
        except KeyError:
            ws.cell(row=row, column=column).value = key
            ws.cell(row=row, column=column + 1).value = item
            row += 1


wb.save(file_name)
print("Cerrando la conexión a la base de datos.")
cursor.close()
cnx.close()
print("Ha finalizado de actualizar los sectores")
input("Presione ENTER para finalizar el programa")
