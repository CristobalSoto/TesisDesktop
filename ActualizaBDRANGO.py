#!/usr/bin/python
# -*- coding: 850 -*-
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from selenium.common.exceptions import NoSuchElementException
from bs4 import BeautifulSoup
import time
# importa la conexi¢n para que esta no sea visible en el c¢digo del programa
from MySQL import cnx
import urllib.request
from openpyxl import load_workbook
import warnings
import os

warnings.filterwarnings("ignore")

anoSem = 2
codRegion = 0
codSector = 0
codTipoProducto = 0
firstloop = True

tabla = "odepa_precios"



def makequery(table, tags, datos, region, producto, sectorp):
    if sectorp == "":
        cueri = ("INSERT INTO " + table +
                 " (region,tipo_producto," + tags + ") "
                 "VALUES ('" + region + "','" + producto + "'," + datos + ")")
    else:
        cueri = ("INSERT INTO " + table +
                 " (region, sector,tipo_producto," + tags + ") "
                 "VALUES ('" + region + "','" + sectorp + "','" + producto + "'," + datos + ")")
    return cueri

time_start = time.time()
cursor = cnx.cursor()
html = urllib.request.urlopen("http://apps.odepa.cl/menu/PreciosConsumidorNS.action?consumidor=frutas&reporte=precios_consumidor")
soup = BeautifulSoup(html, "html.parser")

semana_fin = [str(x.text) for x in soup.find(id="semanaFin").find_all('option')]
regiones = [str(x.text) for x in soup.find(id="codRegion").find_all('option')]
tipos_producto = [str(x.text) for x in soup.find(id="codTipoProducto").find_all('option')]

file_name = os.path.dirname(os.path.abspath(__file__))+'\\01 - CL_CPI_Model.xlsm'
wb = load_workbook(file_name)
ws = wb['Par metros CPI Food']

a¤o_inicio = ws.cell(row=5, column=13).value
semana_inicio = ws.cell(row=5, column=14).value
a¤o_fin = ws.cell(row=6, column=13).value
semana_final = ws.cell(row=6, column=14).value
anoSem = a¤o_inicio

print("Obteniendo los datos de:")
print()
print("        A¤o  Semana")
print("Inicio ", a¤o_inicio, semana_inicio)
print("Termino ", a¤o_fin, semana_final)
#input("Presione ENTER para continuar.")
# carga el browser (chrome, firefox, phantomjs)
chrome_options = Options()
chrome_options.add_argument('--disable-gpu')
driver = webdriver.Chrome(chrome_options=chrome_options)
print("Ingresando datos de:")
for i in range((a¤o_fin-a¤o_inicio)+1):
    for codRegion in range(len(regiones)):
        print()
        print("Region: " + regiones[codRegion])
        for codTipoProducto in range(len(tipos_producto)):
            print("Producto: " + tipos_producto[codTipoProducto])
            driver.get("http://apps.odepa.cl/menu/PreciosConsumidorNS.action?consumidor=frutas&reporte=precios_consumidor")
            # saca los campos del formulario y los llena con el indice
            # mediante el m‚todo "select"
            select = Select(driver.find_element_by_name('params.anoSem'))
            select.select_by_value(str(anoSem))
            select = Select(driver.find_element_by_name('params.semanaIni'))
            select.select_by_index(semana_inicio-1)
            select = Select(driver.find_element_by_name('params.semanaFin'))
            if anoSem == a¤o_fin:
                select.select_by_index(semana_final - 1)
            else:
                select_semanafin = driver.find_element_by_name("params.semanaFin")
                options = [x for x in select_semanafin.find_elements_by_tag_name("option")]
                select.select_by_index(len(options) - 1)

            select = Select(driver.find_element_by_name('params.codRegion'))
            select.select_by_index(codRegion)
            select = Select(driver.find_element_by_name('params.codSector'))
            select.select_by_index(codSector)
            select_sectores = driver.find_element_by_name("params.codSector")
            sectores = [x for x in select_sectores.find_elements_by_tag_name("option")]
            if len(sectores) == 1:
                for obj in sectores:
                    sector = obj.text
            else:
                sector = ""
            select = Select(driver.find_element_by_name('params.codTipoProducto'))
            select.select_by_index(codTipoProducto)

            select = Select(driver.find_element_by_name('params.codProducto'))
            select.select_by_index(0)

            select = Select(driver.find_element_by_name('params.codTipoPuntoMonitoreo'))

            select_monitoreo = driver.find_element_by_name("params.codTipoPuntoMonitoreo")
            options = [x for x in select_monitoreo.find_elements_by_tag_name("option")]

            for element in range(len(options)):
                select.select_by_index(element)
            # enviar formulario
            driver.find_element_by_css_selector('.boton2').click()
            # busca la tabla en el codigo html y obtiene su contenido
            try:
                elem = driver.find_element_by_tag_name('table')
            except NoSuchElementException:
                continue
            contenido = elem.get_attribute('innerHTML')
            contenido = contenido.replace("<br>", " ")
            soup = BeautifulSoup(contenido, "html.parser")
            # busca los tags y los formatea para obtener los datos
            headertags = soup('th')
            count = 0
            columns = ""
            for tag in headertags:
                columns = columns + tag.string.replace(' ', '_').lower() + ","
                count += 1
            columns = columns[:-1]

            info = []
            i = 0
            datatags = soup('td')
            row = "'"
            for tag in datatags:
                if tag.string is None:
                    i += 1
                    row += "','"
                    continue
                if "Elaborado" in tag.string:
                    continue
                if "\'" in tag.string:
                    fila = tag.string.replace('.', '')
                    row = row + str(fila.replace("'", "''")) + "','"
                else:
                    row = row + str(tag.string.replace('.', '')) + "','"
                i += 1
                if i == count:
                    # debug
                    # print(row)
                    row = row[:-2]
                    query = makequery(tabla, columns, row, regiones[codRegion], tipos_producto[codTipoProducto], sector)
                    # debug
                    # print(query)
                    cursor.execute(query)
                    cnx.commit()
                    row = "'"
                    i = 0
    anoSem += 1
    semana_inicio = 1

driver.quit()
print("Ingresando los datos a la tabla.")
cnx.commit()
cursor.execute("UPDATE "+tabla+" "
               "SET producto =  "+tabla+".tipo_producto "
               "WHERE producto Is Null")
cnx.commit()
print("Cerrando la conexi¢n a la base de datos.")
cursor.close()
cnx.close()
duracion = time.time() - time_start
print("Se demoro %s segundos." % duracion)
input("Se ha terminado de actualizar la base de datos, presione ENTER para finalizar.")
