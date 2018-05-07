from openpyxl import load_workbook
from SQLServer import cnx
import os
import warnings
import sys


def makequery(table):
    cueri = ("SELECT DISTINCT(producto), variedad, calidad, procedencia, tipo_producto, unidad FROM "+table+" "
             "WHERE tipo_producto != 'Bebidas' order by tipo_producto")
    return cueri

tabla = "[dbo].[ODEPA_Precios]"
warnings.filterwarnings("ignore")
row = 3
column = 2
print("Cargando la hoja Diccionario del documento excel.")
file_name = '\\\seco\\scan$\\CPI MODEL\\00 - CL_CPI_Food_ODEPA.xlsm'
try:
    wb = load_workbook(file_name, keep_vba=True)
except FileNotFoundError:
    print("No se ha encontrado el archivo.")
    print("Recuerde que el nombre del archivo debe ser '01 - CL_CPI_Model.xlsm' y debe ubicarse en la carpeta Proyecto.")
    input("Presione ENTER para finalizar ")
    sys.exit()
ws = wb['Diccionario ODEPA - INE']

dicc_productos = {}
dicc_productosbd = {}
count = 0
print("Obteniendo los productos desde la hoja Diccionario.")
print("Comparando productos en ambas hojas.")
encontro_diff = True

cursor = cnx.cursor()
query = makequery(tabla)
cursor.execute(query)
# busca hasta que se deje de encontrar diferencias en los productos
while encontro_diff:
    encontro_diff = False
    row = 3
    # se ubica en la fila donde se encuentran los valores
    while ws.cell(row=row, column=column).value is not None:
        tipo_producto = ws.cell(row=row, column=column-1).value
        producto_ine = ws.cell(row=row, column=column+6).value
        producto = ws.cell(row=row, column=column).value
        variedad = ws.cell(row=row, column=column+1).value
        calidad = ws.cell(row=row, column=column+2).value
        if producto_ine == 'CARNE DE PAVO' and ws.cell(row=row, column=column+3).value is None:
            procedencia = ''
        else:
            procedencia = ws.cell(row=row, column=column+3).value
        unidad = ws.cell(row=row, column=column+4).value
        dicc_productos[count] = tipo_producto, producto, variedad, calidad, procedencia, unidad
        count += 1
        row += 1
    count = 0
    # recorre el resultado de la consulta
    for (producto, variedad, calidad, procedencia, tipo_producto, unidad) in cursor:
        dicc_productosbd[count] = tipo_producto, producto, variedad, calidad, procedencia, unidad
        count += 1

    encuentra_producto = False
    # busca por productos que no estén en el documento pero si en la BD
    for key in dicc_productosbd:
        for key2 in dicc_productos:
            if dicc_productosbd[key] == dicc_productos[key2]:
                encuentra_producto = True
                break
        if not encuentra_producto:
            encontro_diff = True
            print("Se ha encontrado el siguiente producto faltante en la hoja Diccionario.")
            print(dicc_productosbd[key])
            print("Se procederá a agregar el producto a hoja de trabajo.")
            input("Presione ENTER para continuar")
            ws.cell(row=row, column=column-1).value = dicc_productosbd[key][0]
            ws.cell(row=row, column=column).value = dicc_productosbd[key][1]
            ws.cell(row=row, column=column+1).value = dicc_productosbd[key][2]
            ws.cell(row=row, column=column+2).value = dicc_productosbd[key][3]
            ws.cell(row=row, column=column+3).value = dicc_productosbd[key][4]
            ws.cell(row=row, column=column+4).value = dicc_productosbd[key][5]
            break
        encuentra_producto = False


wb.save(file_name)
print("Cerrando la conexión a la base de datos.")
cursor.close()
cnx.close()
input("El programa ha terminado de buscar productos nuevos, presione ENTER para salir")
